"""
General procedures for the ONLT

Version 2.0 (2025)

@author: pheijnen
"""

import networkx as nx
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.lines as mlines
import matplotlib.colors as mcolors
import pyvisgraph as vg
import pandas as pd
import numpy as np
import os
import math
from math import inf,ceil,log10
from scipy.spatial.distance import euclidean
from shapely.geometry import Polygon, Point, LineString
from openpyxl import load_workbook,Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from ast import literal_eval
import copy

"""
=============================================================================
PROCEDURE to make a nested dictionary with all the input information
=============================================================================
"""
def input_dict(path, obstacles, existing, routing):
# Get a dicitonary with the demand of all the terminals per timestep
    demand_dict = demand_from_file(path)
# Get a dictionary with all the coordinates of the nodes
    coordinates_dict = coordinates_from_file(path)
# Get a dictionary with all the coordinates of the storage nodes
    coordinates_storage_dict = coordinates_storage_from_file(path)
# Get a dictionary with the amount in storage
    amount_storage_dict = amount_storage_from_file(path)
# Get a dictionary with the capacity off all the storage nodes
    capacity_storage_dict = capacity_storage_from_file(path)
# Get a dictionary with all the obstacles in the network
    obstacles_dict = obstacles_from_file(path, obstacles)
# Get a dictionary with all the existion connections
    existing_connections_dict = existing_connections_from_file(path, existing, routing)
# Get a dictionary of the routing nework
    routing_network_dict = routing_network_from_file(path)
# Make a nested dictionary with all these dictionaries
    input_dict = {
        "demand": demand_dict,
        "coordinates": coordinates_dict,
        "coordinates storage": coordinates_storage_dict,
        "amount": amount_storage_dict,
        "storage": capacity_storage_dict,
        "obstacles": obstacles_dict,
        "existing connections": existing_connections_dict,
        "routing": routing_network_dict
        }
    return input_dict


"""
=============================================================================
PROCEDURE to read coordinates of the nodes from excel-file given by path
=============================================================================
"""
def coordinates_from_file(path):
    try:
        # Read UTM coordinates from the terminals sheet
        data_file = pd.read_excel(path, sheet_name='terminals', usecols=["UTM_Easting", "UTM_Northing"])
        # Construct dictionary of (easting, northing) tuples with integer keys
        coord = {idx: (row["UTM_Easting"], row["UTM_Northing"]) for idx, row in data_file.iterrows()}
    except:
        return {}
    return coord

"""
=============================================================================
PROCEDURES to read storage information from excel-file given by path
=============================================================================
"""
def coordinates_storage_from_file(path):
    try:
# Find number of terminals
        coord_terminals = coordinates_from_file(path)
        numb_terminals = len(coord_terminals)
# Read first sheet of input-file
        data_file = pd.read_excel(path,sheet_name='storage',usecols=["coordinates"])
# Take column with coordinates and transform to dictionary
        data_dict = data_file.iloc[0:,0].to_dict()
# Define dictionary with storage coordinates
        coord_storage = {numb_terminals+node:literal_eval(location)
                         for node,location in data_dict.items()}
        return coord_storage
# If there is no coordinates_storage sheet return an empty dictionary
    except:
        return {}

def amount_storage_from_file(path):
    try:
# Find number of terminals
        coord_terminals = coordinates_from_file(path)
        numb_terminals = len(coord_terminals)
# Read first sheet of input-file
        data_file = pd.read_excel(path,sheet_name='storage',usecols=["amount"])
# Define dictionary with current amount in storage
        data_dict = data_file.iloc[0:,0].to_dict()
        amount_storage = {numb_terminals+node:amount for node,amount in data_dict.items()}
        return amount_storage
# If there is no amount sheet return an empty dictionary
    except:
        return {}

def capacity_storage_from_file(path):
    try:
# Find number of terminals
        coord_terminals = coordinates_from_file(path)
        numb_terminals = len(coord_terminals)
# Read first sheet of input-file
        data_file = pd.read_excel(path,sheet_name='storage',usecols=["capacity"])
# Define dictionary with capacity of storage
        data_dict = data_file.iloc[0:,0].to_dict()
        capacity_storage = {numb_terminals+node:capacity for node,capacity in data_dict.items()}
# If there is no capacity sheet return an empty dictionary
    except:
        return {}
    return capacity_storage

"""
=============================================================================
PROCEDURE to read obstacles from file if any
=============================================================================
"""
def obstacles_from_file(path,obstacles):
    try:
# Start with an empty dictionary of obstacles
        obstacle_list = {}
# If indicated that there are obstacles:
# Read obstacles sheet of input-file
        data_file = pd.read_excel(path,sheet_name='obstacles')
# Loop through all obstacles in the file
        for obstacle in range(data_file.shape[1]):
# Find all corner points
            corner_list = data_file.iloc[:,obstacle].tolist()
            if corner_list:
# Translate obstacle to list
                obstacle_list[obstacle] = [literal_eval(corner)
                                           for corner in corner_list
                                           if not pd.isnull(corner)]
# Return list of obstacles
        return obstacle_list
# If there is no obstacles sheet return an empty dictionary
    except:
        return {}

"""
=============================================================================
PROCEDURE to read existing connections
=============================================================================
"""
def existing_connections_from_file(path,existing,routing):
# Read terminal coordinates from file
    coord = coordinates_from_file(path)
# Read storage information from file
    coord_storage = coordinates_storage_from_file(path)
# Update terminal coordinates with storage
    coord.update(coord_storage)
# If existing connections
    if existing:
# If routing network
        if routing:
# Read routing network
            R = routing_network_from_file(path)
            if R:
# Read coordinates from coordinate network
                coord = nx.get_node_attributes(R,'coord')
        try:
# Read existing connections
            data_file = pd.read_excel(path,sheet_name='existing_connections')
        except:
            existing_edges = {}
            return existing_edges,coord
# Define the dictionary of all existing edges
        existing_edges = {}
# For each existing connection
        for connection in range(data_file.shape[0]):
            
# Read x1, y1, x2, y2, capacity and category
            x1 = data_file.at[connection, 'UTM_Easting_1']
            y1 = data_file.at[connection, 'UTM_Northing_1']
            x2 = data_file.at[connection, 'UTM_Easting_2']
            y2 = data_file.at[connection, 'UTM_Northing_2']
            capacity = data_file.at[connection, 'capacity']
            category = data_file.at[connection, 'Categorie']

            coord1 = (x1, y1)
            coord2 = (x2, y2)

# Add coords to dictionary if not already present
            for coord_point in [coord1, coord2]:
                if coord_point not in coord.values():
                    coord[len(coord)] = coord_point

# Retrieve node keys
            k1 = key_found(coord, coord1)
            k2 = key_found(coord, coord2)

# Store edge with capacity and category
            existing_edges[(k1, k2)] = {
                'capacity': float(capacity),
                'category': category
                }


# If no existing connections
    else:
        existing_edges = {}
# Return existing edges and the total dictionary of coordinates
    return existing_edges,coord

"""
=============================================================================
PROCEDURE to read routing network from file
=============================================================================
"""
def routing_network_from_file(path):
# Read coordinates of terminals from file
    coord = coordinates_from_file(path)
# Determine keys of terminals
    terminals = list(coord.keys())
# Read coordinates of storage from file
    coord_storage = coordinates_storage_from_file(path)
# Determine keys of storage
    storages = list(coord_storage.keys())
# Update dictionary of coordinates with storage coordinates
    coord.update(coord_storage)
    try:
# Read potentional connections in routing network
        data_file = pd.read_excel(path,sheet_name='routing_network')
    except:
        return None
    else:
# Define the list of all edges in routing network
        routing_edges = []
# For each connection in the routing network
        for connection in range(data_file.shape[0]):
# Make list of the coordinates of the end nodes
            Y = data_file.iloc[connection,:].tolist()
# For each coordinate pair in connection
            for P in Y:
# If coordinates are not in the coordinate dictionary yet
                if not literal_eval(P) in coord.values():
# Add coordinates to dictionary and assign to a new node
                    coord[len(coord)] = literal_eval(P)
# Find coordinate keys (= node name) in coordinate dictionary
            k1 = key_found(coord,literal_eval(Y[0]))
            k2 = key_found(coord,literal_eval(Y[1]))
# Add edge between located nodes to edge list
            routing_edges += [(k1,k2)]
        if routing_edges:
# Define the routing network
            G = nx.Graph(routing_edges)
# Set the coordinates as node attributes
            nx.set_node_attributes(G,coord,'coord')
            try:
# Set the label terminal to all terminal nodes
                for node in terminals:
                    G.nodes[node]['stamp'] = 'terminal'
# Set the label storage to all  storage nodes
                for node in storages:
                    G.nodes[node]['stamp'] = 'storage'
# Set the label steiner to all potential steiner nodes in the network
                for node in [node for node in G.nodes() if not node in terminals+storages]:
                    G.nodes[node]['stamp'] = 'steiner'
# Set the length of the connections as weights to the edges
                for node1,node2 in G.edges():
                    G[node1][node2]['weight'] = round(euclidean(coord[node1],coord[node2]),2)
# Return the routing network
                return G
            except:
                print('WARNING: Some terminals or storage nodes are not in routing network.\n Routing network will be ignored \n')
                return None

"""
=============================================================================
PROCEDURE to read demand at all time steps from file
=============================================================================
"""
def demand_from_file(path,k=inf):
    try:
# If all demand time steps are being used
        if k == inf:
# Read 'terminals' sheet of input-file
            data_file = pd.read_excel(path,sheet_name='terminals')
# Else read the specific reduced set of demand time steps
        else:
            data_file = pd.read_excel(path,sheet_name='centres_'+str(k))

# Make a dictionary of dictionaries for each demand time step
        dc = data_file.iloc[:,list(range(6,data_file.shape[1]))].to_dict()
# Make a list of all the time steps as keys
        new_keys=list(range(0,data_file.shape[1]-2))
# Make a list of all the demand values
        new_values = list(dc.values())
# Zip keys and values to demand dictionary of dictionaries
        demand = dict(zip(new_keys,new_values))
# Return demand in all time steps
        return demand
    except:
        return {}

"""
=============================================================================
PROCEDURE to read output network (title) from excel file in path
title = 'MST', 'FT', 'ST', 'bestG'
=============================================================================
"""
def network_from_excel(path,title):
# Read nodes and attributes from file
    df_nodes = pd.read_excel(path,sheet_name=title+'_nodes')
# Read edges and attributes from file
    df_edges = pd.read_excel(path,sheet_name=title+'_edges')
# Make a list of all the nodes in the network
    nodes = [df_nodes['node'][i]
             for i in range(df_nodes.shape[0])]
# Make list of all edges in the network
    edges = [(df_edges['node1'][i],df_edges['node2'][i])
             for i in range(df_edges.shape[0])]
# Define graph from these edges
    G = nx.Graph()
    G.add_edges_from(edges)
    G.add_nodes_from(nodes)
# Set attributes (coordinates, stamp and cost) to nodes in the graph
    for node in range(G.number_of_nodes()):
# Find label of node in network
        old_node = df_nodes['node'][node]
        G.nodes[old_node]['coord']=literal_eval(df_nodes['coord'][node])
        G.nodes[old_node]['stamp']=df_nodes['stamp'][node]
        G.nodes[old_node]['cost']=df_nodes['cost'][node]
# Set attributes, length, category, capacity and current capacity to edges in the graph
    for e in range(df_edges.shape[0]):
# Find labels of end nodes in network
        node1,node2 = df_edges['node1'][e],df_edges['node2'][e]
        G[node1][node2]['weight']=df_edges['length'][e]
        G[node1][node2]['capacity']=df_edges['capacity'][e]
        G[node1][node2]['cost']=df_edges['cost'][e]
        G[node1][node2]['category'] = df_edges['category'][e]
        if df_edges['current'][e] != 0:
            G[node1][node2]['current']=df_edges['current'][e]
# Set graph attribute 'cost'
    cost_edges = sum(nx.get_edge_attributes(G,'cost').values())
    cost_nodes = sum(nx.get_node_attributes(G,'cost').values())
    G.graph['cost'] = cost_edges+cost_nodes
# Return graph with all its attributes
    return G

"""
=============================================================================
PROCEDURE to save network with all attributes to file
=============================================================================
"""
def save_network_to_excel(G,output_path,title):
    G1 = G.copy()
# Define a new empty excel file if new run
    if title == 'MST' or title == 'final network' or not os.path.isfile(output_path):
        wb = Workbook()
        wb.save(output_path)
# Load the workbook
    book = load_workbook(output_path)

# Define dataframe for nodes attributes in graph
    data = ([['node','coord','stamp','cost']]+
            [[node,str(G1.nodes[node]['coord']),
              G1.nodes[node]['stamp'],G1.nodes[node]['cost']]
             for node in G1.nodes()])
    data_file = pd.DataFrame(data)
# Create a new sheet for the nodes
    sheet = book.create_sheet(title+'_nodes')
# Add the nodes and their attributes to the sheet
    for row in dataframe_to_rows(data_file, index=False, header=False):
        sheet.append(row)
# Controleer op edges zonder 'category'
    for node1, node2 in G1.edges():
        if 'category' not in G1[node1][node2]:
            print(f"Edge ({node1}, {node2}) heeft geen 'category' attribuut.")
# Define dataframe for edges and their attributes in graph
    data = [['node1','node2','length','capacity','current','cost','capex', 'opex', 'type','category', 'N_vehicle']]
    for node1,node2 in G1.edges():
# If edge is not an existing connection, set attribute to 0
        if not 'current' in G1[node1][node2]:
            G1[node1][node2]['current']=0
# Add all attributes
        data += [[node1,node2,
                  G1[node1][node2]['weight'],
                  G1[node1][node2]['capacity'],
                  G1[node1][node2]['current'],
                  G1[node1][node2]['cost'],
                  G1[node1][node2]['capex'],
                  G1[node1][node2]['opex'],
                  G1[node1][node2]['new'],
                  G1[node1][node2]['category'],
                  G1[node1][node2]['N_vehicle']
                  ]]
    data_file = pd.DataFrame(data)
# Create a new sheet for the edges
    sheet = book.create_sheet(title+'_edges')
# Add the edges and their attributes to this new sheet
    for row in dataframe_to_rows(data_file, index=False, header=False):
        sheet.append(row)

# Remove empty sheet (if present)
    if 'Sheet' in book.sheetnames:
        std=book['Sheet']
        book.remove(std)
# Save the workbook
    book.save(output_path)

"""
=============================================================================
PROCEDURE to draw obstacles given as list of polygons by corner coordinates
=============================================================================
"""
def draw_obstacles(input_dict):
# Define obstacles from file
    obstacles = input_dict['obstacles']
# Loop through all obstacles
    for obs in obstacles.values():
# Make list of x and y coordinates
        xs, ys = zip(*obs)
# Plot obstacle
        plt.fill(xs,ys, color = 'lightgrey',zorder=0)

"""
=============================================================================
PROCEDURE to draw routing network
=============================================================================
"""
def draw_routing(input_dict,node_size=20):
# Get routing network
    routing_network = input_dict['routing']
# Get coordinates of all nodes in routing network
    coord_routing = nx.get_node_attributes(routing_network,'coord')
# Draw routing network edges
    nx.draw_networkx_edges(routing_network,coord_routing,
                           edge_color='lightgrey')
# Draw routing network nodes
    nx.draw_networkx_nodes(routing_network,coord_routing,
                           node_color='green',node_size=node_size)
# Draw node labels
    if node_size>=100:
        nx.draw_networkx_labels(routing_network,coord_routing,font_size=6,
                                font_color='w')
"""
=============================================================================
PROCEDURE to draw existing connections
=============================================================================
"""
def draw_existing(input_dict, max_width, category_colors):
    edges, pos = input_dict['existing connections']
    pipes = nx.Graph()
    for (n1, n2), attr in edges.items():
        pipes.add_edge(n1, n2, **attr)

    for e, attr in edges.items():
        cap = attr['capacity']
        category = attr.get('category', 'unknown type')
        color = category_colors[category]

        # Buitenranden in kleur
        nx.draw_networkx_edges(
            pipes, pos, edgelist=[e],
            edge_color=color,
            width=float(15 * cap / max_width) + 1  # iets breder dan de grijze
        )

        # Grijs midden
        nx.draw_networkx_edges(
            pipes, pos, edgelist=[e],
            edge_color='lightgrey',
            width=float(15 * cap / max_width)
        )

"""
=============================================================================
PROCEDURE to draw a network G with edge thickness related to assigned capacity
and obstacles or routing network if applicable
=============================================================================
"""
def color_node(typ,deg):
    if typ == 'terminal':
        return 'r'
    elif typ == 'storage':
        return 'purple'
    elif typ == 'steiner':
        return 'green'
    elif typ == 'corner':
        return 'grey'
    elif typ == 'existing':
        if deg == 0:
            return 'lightblue'
        else:
            return 'blue'
    elif typ == 'split':
        return 'green'

def max_edge_width(G):
# Find the max capacity of all edges if any
    if G.edges() and all([('capacity' in G[node1][node2])
                          for node1,node2 in G.edges()]):
        return max(nx.get_edge_attributes(G,'capacity').values())
    else:
        return 1

def draw_network_nodes(G,coord,node_size):
# Make a list of all nodes to keep strict order
    nodes = list(G.nodes())
# Make a list of all node colors based on their stamp
    ncolors = [color_node(G.nodes[i]['stamp'],G.degree(i)) for i in nodes]
# Make a list of all node sizes based on their stamp
    nsize = [node_size if not G.nodes[i]['stamp']
             in ['steiner','split','corner']
             else 0.1*node_size for i in nodes]
# Draw nodes in different colors depending on their stamp
    nx.draw_networkx_nodes(G,coord,nodelist = nodes,
                           node_color=ncolors, node_size = nsize)
# Draw node labels in white if node_size is large enough
    if node_size>=100:
        labels = {i:i for i in G.nodes()
                  if not G.nodes[i]['stamp'] in
                  ['corner']}
        nx.draw_networkx_labels(G,coord,labels=labels,font_size=5,font_color='k')

def draw_network_edges(G,coord,show_capacity,max_width):
# Kleurtoewijzing per transportcategorie
    category_colors = {
        'Pipeline': 'green',
        'Road': 'red',
        'Waterway': 'deepskyblue',
        'unknown type': 'grey'
    }

    # Verdeel edges in bestaande en nieuwe verbindingen
    existing_edges = []
    new_edges = []
    colors_existing = []
    colors_new = []
    widths_existing = []
    widths_new = []

    for node1, node2 in G.edges():
        category = G[node1][node2].get('category', 'unknown type')
        color = category_colors.get(category, 'grey')
        width = float(15 * G[node1][node2]['capacity'] / max_width) if show_capacity else 1
        min_width_new=1
        min_width_existing=1

        if 'current' in G[node1][node2]:  # bestaande verbinding
            existing_edges.append((node1, node2))
            colors_existing.append(color)
            widths_existing.append(max(width, min_width_existing))
        else:  # nieuwe verbinding
            new_edges.append((node1, node2))
            colors_new.append(color)
            widths_new.append(max(width, min_width_new))

    # Teken bestaande verbindingen (volle lijn)
    nx.draw_networkx_edges(G, coord,
                           edgelist=existing_edges,
                           width=widths_existing,
                           edge_color=colors_existing,
                           style='solid')

    # Teken nieuwe verbindingen (stippellijn)
    nx.draw_networkx_edges(G, coord,
                           edgelist=new_edges,
                           width=widths_new,
                           edge_color=colors_new,
                           style='dashed')

    # Voeg legenda toe (transportcategorieën + lijnstijl)
    patches = [mpatches.Patch(color=color, label=label) for label, color in category_colors.items()]
    solid = mlines.Line2D([], [], color='black', linestyle='solid', label='Existing connection')
    dashed = mlines.Line2D([], [], color='black', linestyle='dashed', label='New connection')
    plt.legend(handles=patches + [solid, dashed], loc='lower left', fontsize=8)

def draw_network(G,title,input_dict,routing=False,existing=False,
                 node_size = 100,show_capacity=True, category_colors=None):
    if category_colors is None:
        category_colors = {
            'Pipeline': 'green',
            'Road': 'red',
            'Waterway': 'deepskyblue',
            'Unknown type': 'grey'
        }
    plt.figure(figsize=(20, 20), dpi=1000)
# Find the maximum edge width
    max_width = max_edge_width(G)
# Draw routing network if any
    if routing and input_dict['routing']:
        draw_routing(input_dict,node_size=20)
# Draw existing connections if any
    if existing and input_dict['existing connections']:
        draw_existing(input_dict,max_width,category_colors)
# Draw obstacles
    draw_obstacles(input_dict)
# Get dictionary of coordinates of all nodes in G
    coord = nx.get_node_attributes(G,'coord')
# Draw network nodes
    draw_network_nodes(G,coord,node_size)
# Draw network edges
    draw_network_edges(G,coord,show_capacity,max_width)
# Scale x and y axes
    plt.axis('scaled')
# Print title
    plt.title(title)
    plt.show()

"""
=============================================================================
PROCEDURE to add category of existing connection to new edge from that 
connection
=============================================================================
"""
def inherit_category(allowed_edge_transitions):
    if allowed_edge_transitions and isinstance(allowed_edge_transitions, list):
        return allowed_edge_transitions[0]
    else:
        return 'unknown type'
"""
=============================================================================
PROCEDURE to calculate total cost of network G
based on cost of edges and cost of nodes
=============================================================================
"""
def total_cost(G,beta,spc=0,upc=0,cpc=1, transport_parameters=None, allowed_edge_transitions=None):
# Set cost to all edges
    G1 = edge_cost(G,beta,upc,cpc,transport_parameters=transport_parameters, allowed_edge_transitions=allowed_edge_transitions)
# Set cost to all nodes
    G1 = node_cost(G1,spc)
# Return total cost
    return round(sum(nx.get_edge_attributes(G1,'cost').values()) +
                 sum(nx.get_node_attributes(G1,'cost').values()),0)

"""
=============================================================================
PROCEDURE to calculate cost of each edge in the network G
based on length and capacity of edges with capacity-cost-exponent beta
and if required extra cost for extension of existing capacities
Assign cost as attribute to edges
=============================================================================
"""
def edge_cost(G, beta, upc=0, cpc=1, transport_parameters=None, allowed_edge_transitions=None):
    # Maak een kopie van de graaf
    G1 = G.copy()
    rekenperiode = transport_parameters['Rekenperiode']
    E_GH2 = transport_parameters['E_GH2']
    E_LH2 = transport_parameters['E_LH2']
    for node1, node2 in G1.edges():
        weight = G1[node1][node2]['weight']
        capacity = G1[node1][node2]['capacity']
        category = G1[node1][node2].get('category', 'unknown type')
        #if category == 'unknown type':
            #print(f"Edge {node1}-{node2} heeft geen category — fallback gebruikt")
        #current = G1[node1][node2]['current']

        # Als het een nieuwe verbinding is
        if 'current' not in G1[node1][node2]:
            if beta == 0 and capacity == 0:
                G1[node1][node2]['cost'] = 0
                G1[node1][node2]['category'] = 'unknown type'
                capex = 0
                opex = 0
            else:
                best_cost = float('inf')
                best_type ='unknown type'
                # bereken voor elk edge-type de kosten
                for edge_type in allowed_edge_transitions:
                    if edge_type == 'Pipeline':
                        # Parameters
                        diameter = transport_parameters['Pipeline']['Diameter']
                        behoud_energie = transport_parameters['Pipeline']['Behoud energie']
                        opex_fix_pipeline = transport_parameters['Pipeline']['Opex_fix']
                        
                        #Formules
                        gewicht_H2 = capacity/(E_GH2*behoud_energie)/365
                        capex_full = 50833*weight*math.e**(0.0697*diameter)+297*(diameter**2)+71800*diameter+54658 #function from Hammond (2024), capital costs. Assuming for whole lifetime
                        capex = capex_full/rekenperiode # Divide by calculation period to know costs/year
                        opex = capex*opex_fix_pipeline #opex as percentage of capex
                        
                        cost = capex + opex
                        temp_N_vehicle = None
                    elif edge_type == 'Road':
                        # Parameters
                        cap_truck = transport_parameters['Road']['Cap_truck']
                        cost_var_truck = transport_parameters['Road']['Cost_var']
                        lifetime_truck = transport_parameters['Road']['Lifetime_truck']
                        invest_truck = transport_parameters['Road']['Invest_truck']
                        opex_fix_truck = transport_parameters['Road']['Opex_fix']
                        max_trucks_day = transport_parameters['Road']['max_trucks_per_day']
                        
                        #Formules
                        gewicht_H2 = capacity/E_LH2 #Amount of hydrogen needed in kg for 1 year for this specific edge
                        N_truckloads = math.ceil(gewicht_H2/cap_truck) # Amount of truckloads needed for 1 year, ceiled.
                        
                        N_truck_purchase = math.ceil(N_truckloads/365) #Amount of trucks needed each year to supply demand, assuming a truck can drive every day once to location
                        N_times_new_trucks = rekenperiode/lifetime_truck #Number of times trucks needs to be renewed within rekenperiode
                        
                        capex = (N_truck_purchase * N_times_new_trucks * invest_truck)/rekenperiode
                        
                        opex_fix = capex * opex_fix_truck
                        opex_var = cost_var_truck*weight*2*N_truckloads #Variable opex, assuming a truck needs to drive back to start point as well.
                        opex = opex_fix + opex_var
                        
                        cost = capex + opex
                        temp_N_vehicle = N_truck_purchase
                    else:
                        cost = 9999999  # failsafe
                        capex = 0
                        opex = 0
                    #print(f"Edge {node1}-{node2} | Type: {edge_type} | Cost: {round(cost, 2)}")
                    
                    if cost < best_cost:
                        best_cost = cost
                        best_type = edge_type

                G1[node1][node2]['cost'] = round(best_cost, 2)
                G1[node1][node2]['category'] = best_type
                G1[node1][node2]['N_vehicle'] = temp_N_vehicle

        # === Bestaande verbindingen ===
        else:
            
            if category == 'Pipeline':
                diameter = transport_parameters['Pipeline']['Diameter']
                behoud_energie = transport_parameters['Pipeline']['Behoud energie']
                opex_fix_pipeline = transport_parameters['Pipeline']['Opex_fix']
                
                #Formules
                #gewicht_H2 = capacity/(E_GH2*behoud_energie)/365
                capex_full = 50833*weight*math.e**(0.0697*diameter)+297*(diameter**2)+71800*diameter+54658
                capex = capex_full/rekenperiode
                opex = capex*opex_fix_pipeline
                
                G1[node1][node2]['cost'] = round(opex, 2)
                G1[node1][node2]['N_vehicle'] = None
            elif category == 'Road':
                # Parameters
                cap_truck = transport_parameters['Road']['Cap_truck']
                cost_var_truck_snelweg = transport_parameters['Road']['Cost_var_snelweg']
                lifetime_truck = transport_parameters['Road']['Lifetime_truck']
                invest_truck = transport_parameters['Road']['Invest_truck']
                opex_fix_truck = transport_parameters['Road']['Opex_fix']
                max_trucks_day = transport_parameters['Road']['max_trucks_per_day']
                
                #Formules
                gewicht_H2 = capacity/E_LH2
                N_truckloads = math.ceil(gewicht_H2/cap_truck)
                
                N_truck_purchase = math.ceil(N_truckloads/365) #Amount of trucks needed each year to supply demand
                N_times_new_trucks = rekenperiode/lifetime_truck #Number of times trucks needs to be renewed within rekenperiode
                
                capex = (N_truck_purchase * N_times_new_trucks * invest_truck)/rekenperiode
                
                opex_fix = capex * opex_fix_truck
                opex_var = cost_var_truck_snelweg*weight*2*N_truckloads
                opex = opex_fix + opex_var
                
                cost = opex #Only opex costs, capex costs are accounted for in new edges, because we assume it is the same truck which drives on a new edge and
                G1[node1][node2]['cost'] = round(cost, 2)
                G1[node1][node2]['N_vehicle'] = N_truck_purchase
            elif category == 'Waterway':
                # Parameters
                cap_boat = transport_parameters['Waterway']['Cap_boot']
                cost_var_boat = transport_parameters['Waterway']['Cost_var']
                lifetime_boat = transport_parameters['Waterway']['Lifetime_boot']
                invest_boat = transport_parameters['Waterway']['Invest_boot']
                opex_fix_boat = transport_parameters['Waterway']['Opex_fix']
                
                # Formules
                gewicht_H2 = capacity/E_LH2/365
                N_boatloads = math.ceil(gewicht_H2/cap_boat)
                
                N_boat_purchase = math.ceil(N_boatloads/365) #Amount of trucks needed each year to supply demand
                N_times_new_boats = rekenperiode/lifetime_boat #Number of times trucks needs to be renewed within rekenperiode
                
                capex = (N_boat_purchase * N_times_new_boats * invest_boat)/rekenperiode
                
                opex_fix = capex * opex_fix_boat
                opex_var = cost_var_boat*weight*2*N_boatloads
                opex = opex_fix + opex_var
                
                cost = capex + opex 
                G1[node1][node2]['cost'] = round(cost, 2)
                G1[node1][node2]['N_vehicle'] = N_boat_purchase
            else:
                G1[node1][node2]['cost'] = 9999999  # failsafe
                capex = 0
                opex = 0
    
        G1[node1][node2]['capex'] = round(capex, 2)
        G1[node1][node2]['opex'] = round(opex, 2)
        G1[node1][node2]['new'] = 'New' if 'current' not in G1[node1][node2] else 'Existing'

    return G1


"""
=============================================================================
PROCEDURE to calculate cost of each node in the network G
only for splitting and stiner nodes (spc)
Assign cost as attribute to nodes
=============================================================================
"""
def node_cost(G,spc=0):
# Make a copy of the graph
    G1 = G.copy()
# For each node in the graph
    for node in G1.nodes():
# If node is a steiner node or split node and adjacent edges have capacity
        if (G1.nodes[node]['stamp'] in ['steiner','split'] and
            sum(G1[i][j]['capacity'] for i,j in G1.edges(node))>0):
# Add cost given by spc
            G1.nodes[node]['cost'] = spc
# For all other nodes
        else:
# Add no cost
            G1.nodes[node]['cost'] = 0
# Return graph with cost attribute to nodes
    return G1

"""
=============================================================================
PROCEDURE to extend the graph G with artifical nodes
if total demand <> total supply (td <> 0)
=============================================================================
"""
def extend_graph(G,demand,scaling_factor):
# Make a copy of the graph
    G1 = G.copy()
# Add super producer (sp) and super consumer (sc)
    G1.add_nodes_from(['sp','sc'])
# For each non storage node
    for node in [node for node in G.nodes()
                 if G.nodes[node]['stamp']!='storage']:
# If node is a consumer in this time step
        if node in demand.keys() and demand[node]>0:
# Add a no-cost edge between node an super consumer with capacity equal to node demand
            G1.add_edge(node,'sc',
                        capacity=round(demand[node]*scaling_factor),
                        weight=0)
# If node is a producer in this time step
        if node in demand.keys() and demand[node]<0:
# Add a no-cost edge between super producer and node with capacity equal to node supply
            G1.add_edge('sp',node,
                        capacity=-round(demand[node]*scaling_factor),
                        weight=0)
# For each storage node
    for storage in [node for node in G.nodes()
                    if G.nodes[node][ 'stamp']=='storage']:
# Add a no-cost edge between super producer and storage with capacity equal to amount in storage
        G1.add_edge('sp',storage,
                    capacity=round(G.nodes[storage]['amount']*
                                   scaling_factor),
                    weight=0)
# Add a very expensive edge between storage and super consumer with capacity
# equal to capacity of storage in order to avoid using this edge and first supply terminals
        G1.add_edge(storage,'sc',
                    capacity=round(G.nodes[storage]['capacity']*
                                   scaling_factor),
                    weight=100000000)
# Return graph with extra nodes and edges
    return G1


"""
=============================================================================
PROCEDURE to assign capacity to edges in a graph G
based on the given demand-supply pattern demand for one time step
=============================================================================
"""

def capacity_graph(G,demand,scaling_factor=1):
# Make a copy of the graph
    G1 = G.copy()
# Round edge weights and set infinite capacity in order to use max_flow_min_cost
    for node1,node2 in G1.edges():
        G1[node1][node2]['weight']=round(G1[node1][node2]['weight'])
        G1[node1][node2]['capacity']=inf
# Make graph directed
    G1 = G1.to_directed()
# Extend the graph with super source and super sink
    G1 = extend_graph(G1,demand,scaling_factor)
# Apply max flow min cost to find minimal required capacity on edges
    mf = nx.max_flow_min_cost(G1,'sp','sc')
# For each storage node
    for storage in [node for node in G.nodes() if G.nodes[node]['stamp']=='storage']:
# Add to amount what goes to super consumer
        G1.nodes[storage]['amount']+=mf[storage]['sc']/scaling_factor
# Subtract from amount what comes from super producer
        G1.nodes[storage]['amount']-=mf['sp'][storage]/scaling_factor
# Remove super consumer and producer from graph
    G1.remove_nodes_from(['sp','sc'])
# For each edge in the graph
    for node1,node2 in G1.edges():
# Set edge capacity
        G1[node1][node2]['capacity']=(max(mf[node1][node2],mf[node2][node1])
                                      /scaling_factor)
# Return the graph with required capacity as edge attribute
    return G1

"""
=============================================================================
PROCEDURE to assign max needed capacity to graph G
over all demand-supply profiles in path
=============================================================================
"""
def final_capacity_graph(input_dict,deqs,G):
# If demand equals supply and existing==False use leaf_assignment to calculate the capacities
    if deqs:
        return leaf_assignment(input_dict, G)
# Make copy of graph to update capacity
    G_final = G.copy()
# Set edge capacities to 0
    nx.set_edge_attributes(G_final,0,'capacity')
### AANGEPAST
# For each existing connection
   # for node1,node2 in [(i,j) for i,j in G_final.edges() if 'current' in G_final[i][j]]:
# Set capacity to current capacity
    #    G_final[node1][node2]['capacity'] = G_final[node1][node2]['current']
# For all storage nodes get amount in storage
    amount_storage = input_dict['amount']
# For all storage nodes get capacity of storage
    capacity_storage = input_dict['storage']
# For all storage nodes
    for storage in [node for node in G.nodes() if G.nodes[node]['stamp']=='storage']:
# Set amount in storage
        G.nodes[storage]['amount']=amount_storage[storage]
# Set storage capacity
        G.nodes[storage]['capacity']=capacity_storage[storage]

# Read sheet of input-file with demand over all time steps
    demand_all = input_dict['demand']
# Find max value in all demand
    max_dem=max([abs(i) for j in demand_all.keys()
           for i in demand_all[j].values()])
# Find scaling factor to round off capacity values
    scaling_factor = max(1,10**(4-ceil(log10(max_dem))))
# Make a copy of the graph
    G1 = G.copy()
# For each demand-supply pattern
    for demand in demand_all.values():
# Set minimal edge capacity to satisfy demand
        G1 = capacity_graph(G1,demand,scaling_factor)
# For each edge in the graph
        for node1,node2 in G1.edges():
# Update capacity in final graph if larger capacity needed
            G_final[node1][node2]['capacity'] = max(G_final[node1][node2]['capacity'],
                                                    G1[node1][node2]['capacity'])
# Return final graph with assigned capacity to the edges
    return G_final

"""
=============================================================================
PROCEDURE to assign capacity to edges in a graph G
based on the given demand-supply pattern demand for one time step when
=============================================================================
"""
def leaf_assignment(input_dict, G):
# Copy of graph to update capacities
    G1 = G.copy()
# Set all edge capacities to 0
    nx.set_edge_attributes(G1,0,'capacity')
# Iterate through all the timesteps
    for timestep, demand in input_dict['demand'].items():
# Copy of graph to remove edges
        T = G.copy()
# Set demand 0 for all nodes not given in demand-supply pattern
        demand.update({j:0 for j in [i for i in G.nodes()
                                     if not i in demand.keys()]})
# Copy demand to update demand-supply left
        Q = copy.deepcopy(demand)
# While not all edges in T scanned
        while T.edges():
# Find leaves of the graph
            L = [s for s in T.nodes() if T.degree[s]==1]
# While leaves in the list
            while L:
# Take one
                i = L.pop()
# Find neighbor of leaf
                B = list(T.neighbors(i))
# If neighbor found
                if B:
# Update the capacity if it is higher then the already assigned capacity
                    if abs(Q[i]) > G1[i][B[0]]['capacity']:
                        G1[i][B[0]]['capacity'] = abs(Q[i])
# Update left demand or supply for neighbor node
                    Q[B[0]] = Q[B[0]]+Q[i]
# Remove edge from T
                    T.remove_edge(i,B[0])
# Return capacitated graph
    return G1


"""
=============================================================================
PROCEDURE to remove obsolete nodes in G
=============================================================================
"""
def remove_obsolete_nodes(G,routing=False):
# Find all nodes that have only zero-capacity edges
    nodes_to_remove = [i for i in G.nodes()
                       if (sum([G[i][j]['capacity'] for j in G.neighbors(i)])==0 or
                           G.degree(i)<=1)
                       and not (G.nodes[i]['stamp'] in ['terminal','storage','existing'])]
# Copy graph G
    G1 = G.copy()
# Remove obsolete nodes
    G1.remove_nodes_from(nodes_to_remove)
# Relabel nodes only if not routing
    if not routing:
# Return relabeled graph
        return relabeled_graph(G1)
    return G1

"""
=============================================================================
PROCEDURE to add weights (lengths = euclidean distance) to edges in G
=============================================================================
"""
def euclidean_weighted_graph(G):
# Make copy of graph
    G1 = G.copy()
# Loop through all edges
    for node1,node2 in G1.edges():
# Add euclidean distance as weight to edge
        G1[node1][node2]['weight'] = euclidean(G1.nodes[node1]['coord'],
                                               G1.nodes[node2]['coord'])
# Return weighted graph G1
    return G1

"""
=============================================================================
Procedure to define visibility graph from obstacles
=============================================================================
"""
def visibility_graph(dict_obstacles):
# Define polygons for each obstacle
    polys = [[vg.Point(node1,node2) for node1,node2 in obs]
             for obs in dict_obstacles.values()]
# Define visibility graph
    G = vg.VisGraph()
# Build visibility around obstacles
    G.build(polys,status=False)
# Return visibility graph
    return G

"""
=============================================================================
Procedure to check if edge from coordinates a to b crosses obstacle
=============================================================================
"""
def crosses_obstacle(a,b,obstacles):
# Set boolean to false
    cross = False
# Define line segment between coordinate points a and b
    line = LineString([a,b])
# Loop through all obstacles
    for obs in obstacles.values():
# Define polygon for each obstacle
        polygon = Polygon(obs)
# Check if line crosses polygon
        if line.crosses(polygon):
# Set cross to true
            cross = True
# Break the for-loop
            break
# Return boolean
    return cross

"""
=============================================================================
Procedure to find the nearest polygon corner for a point in an obstacle
=============================================================================
"""
def nearest_polygon_vertex(vertex,dict_obstacles):
# Copy point
    vertex1 = vertex
# For each obstacle
    for obstacle in dict_obstacles.values():
# Define a polygon for the obstacle
        polygon = Polygon(obstacle)
# Define point for p
        point = Point(*vertex)
# If point p lies within polygon
        if polygon.contains(point):
# Calculate dictionary with distance to each corner of polygon
            dist = [(corner,euclidean(vertex,corner)) for corner in obstacle]
# Find point with smallest distance to p
            vertex1 = min(dist, key=lambda x: x[1])[0]
            break
# Return new point
    return vertex1

"""
=============================================================================
PROCEDURE to redirect edges of graph G around obstacles
=============================================================================
"""
def redirected_graph(G,input_dict,obstacles):
# Find obstacles if any
    dict_obstacles = input_dict['obstacles']
# Make a copy of the graph
    G1 = G.copy()
# If there are obstacles
    if obstacles:
# Define visibility graph from obstacles
        visg = visibility_graph(dict_obstacles)
# Shift nodes to outside obstacles
        G1 = shifted_nodes(G1,dict_obstacles)
# Define list of all edges of G that are not existing connections
        non_exist_edges = [(node1,node2) for node1,node2 in G1.edges()
                           if not 'current' in G1[node1][node2]]
# For each non-existing connection
        for node1,node2 in non_exist_edges:
# Redirect edge around obstacles
            G1 = redirected_edge(node1,node2,G1,visg)
# Remove Steiner and splitting nodes no longer needed
        G1 = cleaned_split_nodes(G1)
# Relabel the graph nodes from 0 to number of nodes
        G1 = relabeled_graph(G1)
# Add weights to all edges of the graph
        G1 = euclidean_weighted_graph(G1)
# Remove cycles in the graph
        G1 = removed_cycles(G1)
# Return the redirected graph
    return G1

"""
=============================================================================
PROCEDURE to redirect edge (i,j) in graph
to shortest path in visibility graph
=============================================================================
"""
def redirected_edge(node1,node2,G,visg):
# Make copy of graph
    G1 = G.copy()
# Find coordinates of all nodes in the graph
    coord = nx.get_node_attributes(G1,'coord')
# Find total number of nodes in the graph
    total_nodes = G1.number_of_nodes()
# Determine shortest path around obstacles connecting end nodes
    short_path = visg.shortest_path(vg.Point(*coord[node1]),vg.Point(*coord[node2]))
# Make list of node coordinates on shortest path
    sp_coord = [(s.x,s.y) for s in short_path]
# If path is longer than 2 nodes then edge is redirected
    if len(sp_coord)>2:
# Loop through all coordinate points on shortest path
        for node,location in enumerate(sp_coord[1:],start=1):
# If coordinate points is not in graph yet
            if not location in coord.values():
# Add new coordinate point to coordinates dictionary
                coord[total_nodes] = location
# Add new node to graph
                G1.add_node(total_nodes,stamp='corner',coord=coord[total_nodes])
# Update total number of nodes in graph
                total_nodes += 1
# Find node numbers in coordinate dictionary for edge on shortest path
            k1 = key_found(coord,sp_coord[node-1])
            k2 = key_found(coord,location)
# Add edge between found nodes
            G1.add_edge(k1,k2,weight = euclidean(coord[k1],coord[k2]))
# Remove original edge between end nodes
        G1.remove_edge(node1,node2)
# Return graph with redirected edge
    return G1

"""
=============================================================================
PROCEDURE to find key in dictionary dc for a given value val
=============================================================================
"""
def key_found(dc,val):
# Find key of given value in the dictionary
    return next((key for key, value in dc.items() if val == value), None)

"""
=============================================================================
PROCEDURE to shift nodes in graph G to nearest ostacle corner
=============================================================================
"""
def shifted_nodes(G,dict_obstacles):
# Make copy of graph
    G1 = G.copy()
# For each node in the graph
    for node in G1.nodes():
# If node is not end node of existing connection
        if not any(['current' in G1[node1][node2] for node1,node2 in G1.edges(node)]):
# Shift node to the nearest polygon vertex
            G1.nodes[node]['coord'] = nearest_polygon_vertex(G1.nodes[node]['coord'],
                                                             dict_obstacles)
# Return graph with shifted nodes
    return G1

"""
=============================================================================
PROCEDURE to remove Steiner en splitting nodes in graph G
with degree smaller than 2
=============================================================================
"""
def cleaned_split_nodes(G):
# Make copy of graph
    G1 = G.copy()
# Find all Steiner and splitting nodes that are no longer needed
    nodes_to_remove = [node for node in G1.nodes()
                       if G1.nodes[node]['stamp'] in ['steiner','split'] and
                       G1.degree(node) <= 1]
# Remove all unneeded nodes
    G1.remove_nodes_from(nodes_to_remove)
# Return cleaned graph
    return G1

"""
=============================================================================
PROCEDURE to remove cycles in graph G
by removing the longest edge in the cycle
=============================================================================
"""
def removed_cycles(G):
# Make a copy of the graph
    G1 = G.copy()
# As long as there are cycles in the graph
    cycle = not_existing_cycle(G1)
# If cyle found
    if cycle:
# Define dictionary of all edge weights for all edges in cycle that are not existing connections
        W = {(node1,node2): G1[node1][node2]['weight']
             for (node1,node2) in zip(cycle[:-1],cycle[1:])
             if not 'current' in G1[node1][node2]}
# Sort edges on their weight from small to large
        W_sort = sorted(W.items(), key=lambda x: x[1])
# Take the longest edge from the cycle
        (node1,node2),length = W_sort.pop()
# Remove longest edge from graph from the graph
        if (node1,node2) in G1.edges():
            G1.remove_edge(node1,node2)
        else:
            G1.remove_edge(node2,node1)
# Return the graph without cycles
    return G1

"""
=============================================================================
PROCEDURE to find a cycle in graph G
with not all edges existing connections
=============================================================================
"""
def not_existing_cycle(G):
# Find cycle basis
    cb = nx.cycle_basis(G)
# Initialize empty cycle
    nc = []
# For every cycle
    for c in cb:
# Make it a closed cycle
        c += [c[0]]
# Check if not every edge in the cycle is an existing connection
        if not all(['current' in G[i][j] for i,j in zip(c[:-1],c[1:])]):
            nc = c
            break
# Return the cycle found
    return nc

"""
=============================================================================
PROCEDURE to relabel nodes in graph from 0 to total number of nodes minus 1
=============================================================================
"""
def relabeled_graph(G):
# Gather all different type of nodes
    terminals = [i for i in G.nodes() if G.nodes[i]['stamp']=='terminal']
    storages = [i for i in G.nodes() if G.nodes[i]['stamp']=='storage']
    corners = [i for i in G.nodes() if G.nodes[i]['stamp'] == 'corner']
    steiner = [i for i in G.nodes() if G.nodes[i]['stamp'] == 'steiner']
    splits = [i for i in G.nodes() if G.nodes[i]['stamp'] == 'split']
    existing = [i for i in G.nodes() if G.nodes[i]['stamp'] == 'existing']
# Define mapping dictionary in which terminals,storages and existing nodes remain the same
    mapping = {j:j for j in terminals+storages+existing}
# Find next node number
    new_node = len(terminals+storages+existing)
# For each node in the remaining list of nodes
    for old_node in corners+splits+steiner:
# Assign next number to next node
        mapping[old_node] = new_node
# Update new_node number
        new_node += 1
# Relabel all nodes in G
    G1 = nx.relabel_nodes(G,mapping)
# Return G1
    return G1

"""
=============================================================================
PROCEDURE to find best forest in case of a cycle
=============================================================================
"""
def best_forest(deqs,G,input_dict,routing,beta,spc,upc,cpc,transport_parameters,allowed_edge_transitions):
# Find all existing edges in graph
    existing_edges = [(node1,node2) for node1,node2 in G.edges()
                      if 'current' in G[node1][node2]]
# Copy graph
    bestG = G.copy()
# Set initial cost infinitely high
    bestG.graph['cost'] = inf
# Find cycle
    for cb in nx.cycle_basis(G):
# Close cycle
        cb += [cb[0]]
# Remove each edge in cycle one by one if not an existing connection
        for node1,node2 in [(i,j) for i,j in list(zip(cb[0:],cb[1:]))
                            if not 'current' in G[i][j]]:
# Copy the current network with cycle
            G2 = G.copy()
# Remove edge
            if (node1,node2) in G2.edges():
                G2.remove_edge(node1,node2)
            else:
                G2.remove_edge(node2,node1)
# Remove existing connections from graph
            FG = G2.copy()
            FG.remove_edges_from(existing_edges)
# If remaining network is a forest
            if nx.is_forest(FG):
# Assign final capacity based on k demand-supply profiles
                G2 = final_capacity_graph(input_dict, deqs,G2)
# Calculate new cost
                G2.graph['cost'] = total_cost(G2,beta,spc,upc,cpc,transport_parameters=transport_parameters, 
                                              allowed_edge_transitions=allowed_edge_transitions)
# If cheaper network
                if G2.graph['cost'] < bestG.graph['cost']:
# Store best network
                    bestG = G2.copy()
# Return
    return bestG

"""
=============================================================================
PROCEDURE to determine maximum flow through network with given capacity and
demand-supply (not taking into account storage nodes)
=============================================================================
"""
def max_flow(G,demand):
# Make a copy of the graph
    G2 = G.copy()
# Add attribute 'demand' with value 0 to all nodes
    nx.set_node_attributes(G2,0,'demand')
# Update attribute 'demand' for supply or consumption nodes
    nx.set_node_attributes(G2,demand,'demand')
# Make directed graph
    G1 = G2.to_directed()
# Add super producer (sp) and super consumer (sc)
    G1.add_nodes_from(['sp','sc'])
# Add edges from super producer to all supply nodes with capacity equal to supply
    G1.add_edges_from([('sp',node,{'capacity':-G2.nodes[node]['demand']})
                       for node in G2.nodes() if G2.nodes[node]['demand']<0])
# Add edges from consumer nodes to super consumer with capacity equal to demand
    G1.add_edges_from([(node,'sc',{'capacity':G2.nodes[node]['demand']})
                       for node in G2.nodes() if G2.nodes[node]['demand']>0])
# Calculate the maximum flow through the network from super producer to super consumer
    max_flow,edge_flows = nx.maximum_flow(G1,'sp','sc')
# Return the maximum flow
    return max_flow

"""
=============================================================================
PROCEDURE to plot the storage levels over time for a given network and demand-supply time steps
=============================================================================
"""
def plot_storage(G,input_dict, title):
# Find all storage nodes in the graph
    storages= [node for node in G.nodes()
               if G.nodes[node]['stamp']=='storage']

# Define dictionary with initial amount in storage
    amount_storage = input_dict['amount']
# Define dictionary with capacity of storage nodes
    capacity_storage = input_dict['storage']
# Define dictionary for amount of storage nodes over all time steps
    amount_in_storages = {}

# For each storage node
    for node in storages:
# Set amount in storage to initial state
        G.nodes[node]['amount']=amount_storage[node]
# Set capacity to storage capacity
        G.nodes[node]['capacity']=capacity_storage[node]
# Update amount in storage for time step 0
        amount_in_storages[node] = [amount_storage[node]]

# Make a copy of the original graph
    G1 = G.copy()
# For each demand-supply pattern
    for demand in input_dict['demand'].values():
# Assign capacity to new graph G1 for this demand-supply pattern
        G1 = capacity_graph(G1,demand)
# For each storage node
        for node in storages:
# Update amount in storage
            amount_in_storages[node]+=[G1.nodes[node]['amount']]
# For each storage node
    for node in storages:
# Plot amount in storage for each time step
        plt.plot(list(range(len(amount_in_storages[node]))),amount_in_storages[node],
                 label='storage_'+str(node))
# Ticks on horizontal axis
    plt.xticks(np.arange(0, len(amount_in_storages[storages[0]]), 1))
# Ax labels
    plt.xlabel("Time step")
    plt.ylabel("Amount in storage")
    plt.title(f"{title}")
# Plot legend
    plt.legend()
    plt.show()

"""
=============================================================================
PROCEDURE to check if there is storage in network
=============================================================================
"""
def storage_in_network(inputfile):
    try:
# Load the Excel file (assuming the file is named 'file.xlsx' and on the first sheet)
        df = pd.read_excel(inputfile, sheet_name='storage')  # Adjust the sheet name if needed

# Specify the column you want to check
        column_name = 'coordinates'  # Replace 'ColumnName' with the actual column name

# Check if the column is empty (contains only NaN values)
        is_empty = df[column_name].isna().all()
# If there is no storage return false
        if is_empty:
            storage = False
# Else return true
        else:
            storage = True
        return storage
# Return false if the whole worksheet doesn't exist
    except:
        return False

"""
=============================================================================
PROCEDURE to check if the demand equals supply in every timestep
=============================================================================
"""
def demand_equals_supply(demand_dict):
# Initialize a dictionary to hold the sums of each key
    sums = {}

# Iterate through each inner dictionary and add values to the corresponding key in the `sums` dictionary
    for timestep, demand in demand_dict.items():
        sums[timestep] = 0
        for value in demand.values():
            sums[timestep] += value

# Check if all sums are zero
    all_zero = all(total == 0 for total in sums.values())
# Return true if all the sums are zero, otherwise return false
    return all_zero
